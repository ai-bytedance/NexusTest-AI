from __future__ import annotations

import csv
import io
from typing import Any, Iterable
from uuid import UUID

from fastapi import APIRouter, Depends, Request, UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.response import ResponseEnvelope, success_response
from app.core.authz import ProjectContext, require_project_member
from app.core.errors import ErrorCode, http_exception
from app.db.session import get_db
from app.models.dataset import Dataset, DatasetType
from app.schemas.dataset import DatasetCreate, DatasetPreview, DatasetRead, DatasetUpdate
from app.services.datasets.loader import DatasetLoadError, load_dataset_rows
from app.services.datasets.storage import DatasetStorage

router = APIRouter(prefix="/projects/{project_id}/datasets", tags=["datasets"])


def _serialize(dataset: Dataset) -> DatasetRead:
    payload = {
        "id": dataset.id,
        "project_id": dataset.project_id,
        "name": dataset.name,
        "type": dataset.type,
        "schema": dataset.schema or {},
        "source": dataset.source or {},
        "created_by": dataset.created_by,
        "created_at": dataset.created_at,
        "updated_at": dataset.updated_at,
        "is_deleted": dataset.is_deleted,
    }
    return DatasetRead.model_validate(payload)


def _dataset_storage() -> DatasetStorage:
    return DatasetStorage()


def _get_dataset(db: Session, project_id: UUID, dataset_id: UUID) -> Dataset:
    stmt = (
        select(Dataset)
        .where(
            Dataset.id == dataset_id,
            Dataset.project_id == project_id,
            Dataset.is_deleted.is_(False),
        )
        .limit(1)
    )
    dataset = db.execute(stmt).scalar_one_or_none()
    if dataset is None:
        raise http_exception(status.HTTP_404_NOT_FOUND, ErrorCode.NOT_FOUND, "Dataset not found")
    return dataset


def _extract_columns_from_rows(rows: Iterable[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        for key in row.keys():
            if isinstance(key, str) and key not in columns:
                columns.append(key)
        if columns:
            break
    return columns


def _build_schema(columns: list[str]) -> dict[str, Any]:
    return {"columns": columns}


def _parse_inline_payload(payload: DatasetCreate) -> tuple[dict[str, Any], list[str]]:
    source = payload.source or {}
    rows = source.get("rows")
    if not isinstance(rows, list) or not rows:
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Inline dataset requires rows")
    if any(not isinstance(row, dict) for row in rows):
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Each inline row must be an object")
    columns = _extract_columns_from_rows(rows)
    if not columns:
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Unable to infer dataset columns")
    return {"rows": rows}, columns


def _compute_csv_columns(file_bytes: bytes) -> list[str]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "CSV file must be UTF-8 encoded") from exc
    buffer = io.StringIO(text)
    reader = csv.reader(buffer)
    try:
        header = next(reader)
    except StopIteration as exc:
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "CSV file is empty") from exc
    columns = [column.strip() for column in header if column.strip()]
    if not columns:
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "CSV header row is empty")
    return columns


def _compute_excel_columns(file_bytes: bytes, *, sheet: str | None = None) -> tuple[list[str], str | None]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, f"Sheet '{sheet}' not found")
            worksheet = workbook[sheet]
            resolved_sheet = sheet
        else:
            worksheet = workbook.active
            resolved_sheet = worksheet.title
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:  # pragma: no cover - empty worksheet
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Excel worksheet is empty") from exc
        columns = [str(cell).strip() for cell in header_row if cell is not None and str(cell).strip()]
        if not columns:
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Excel header row is empty")
        return columns, resolved_sheet
    finally:
        workbook.close()


def _validate_dataset_name(name: str | None) -> str:
    if not name or not name.strip():
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Dataset name is required")
    return name.strip()


@router.get("", response_model=ResponseEnvelope)
def list_datasets(
    context: ProjectContext = Depends(require_project_member),
    db: Session = Depends(get_db),
) -> dict:
    stmt = (
        select(Dataset)
        .where(
            Dataset.project_id == context.project.id,
            Dataset.is_deleted.is_(False),
        )
        .order_by(Dataset.created_at.asc())
    )
    datasets = db.execute(stmt).scalars().all()
    data = [_serialize(item) for item in datasets]
    return success_response(data)


@router.post("", response_model=ResponseEnvelope, status_code=status.HTTP_201_CREATED)
async def create_dataset(
    request: Request,
    context: ProjectContext = Depends(require_project_member),
    db: Session = Depends(get_db),
) -> dict:
    content_type = request.headers.get("content-type", "")
    storage = _dataset_storage()

    if "multipart/form-data" in content_type:
        form = await request.form()
        name = _validate_dataset_name(form.get("name"))
        dataset_type_raw = (form.get("type") or "").strip().lower()
        try:
            dataset_type = DatasetType(dataset_type_raw)
        except ValueError as exc:
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Invalid dataset type") from exc
        upload = form.get("file")
        if not isinstance(upload, UploadFile):
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Dataset file is required")
        file_bytes = await upload.read()
        if not file_bytes:
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Uploaded file is empty")

        if dataset_type == DatasetType.CSV:
            columns = _compute_csv_columns(file_bytes)
            sheet_name = None
        elif dataset_type == DatasetType.EXCEL:
            sheet_param = form.get("sheet")
            columns, sheet_name = _compute_excel_columns(file_bytes, sheet=str(sheet_param) if sheet_param else None)
        else:
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Inline datasets must be uploaded via JSON")

        dataset = Dataset(
            project_id=context.project.id,
            name=name,
            type=dataset_type,
            schema=_build_schema(columns),
            source={},
            created_by=context.membership.user_id,
        )
        db.add(dataset)
        db.flush()

        relative_path = storage.save_file(
            str(context.project.id),
            str(dataset.id),
            upload.filename or f"dataset.{dataset_type.value}",
            file_bytes,
        )
        source_payload: dict[str, Any] = {
            "path": relative_path,
            "filename": upload.filename or relative_path.rsplit("/", 1)[-1],
        }
        if sheet_name:
            source_payload["sheet"] = sheet_name
        dataset.source = source_payload

        db.add(dataset)
        db.commit()
        db.refresh(dataset)
        return success_response(_serialize(dataset))

    raw_payload = await request.json()
    if not isinstance(raw_payload, dict):
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Invalid JSON payload")
    payload = DatasetCreate.model_validate(raw_payload)
    if payload.type != DatasetType.INLINE:
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "File datasets must be uploaded via multipart form")
    source_payload, columns = _parse_inline_payload(payload)

    dataset = Dataset(
        project_id=context.project.id,
        name=_validate_dataset_name(payload.name),
        type=DatasetType.INLINE,
        schema=_build_schema(columns),
        source=source_payload,
        created_by=context.membership.user_id,
    )
    db.add(dataset)
    db.commit()
    db.refresh(dataset)
    return success_response(_serialize(dataset))


@router.get("/{dataset_id}", response_model=ResponseEnvelope)
def get_dataset(
    dataset_id: UUID,
    context: ProjectContext = Depends(require_project_member),
    db: Session = Depends(get_db),
) -> dict:
    dataset = _get_dataset(db, context.project.id, dataset_id)
    return success_response(_serialize(dataset))


@router.patch("/{dataset_id}", response_model=ResponseEnvelope)
async def update_dataset(
    dataset_id: UUID,
    request: Request,
    context: ProjectContext = Depends(require_project_member),
    db: Session = Depends(get_db),
) -> dict:
    dataset = _get_dataset(db, context.project.id, dataset_id)
    storage = _dataset_storage()
    content_type = request.headers.get("content-type", "")

    if "multipart/form-data" in content_type:
        form = await request.form()
        name_value = form.get("name")
        if name_value is not None:
            dataset.name = _validate_dataset_name(str(name_value))
        type_value = form.get("type")
        upload = form.get("file")
        sheet_param = form.get("sheet")

        if upload is not None:
            if not isinstance(upload, UploadFile):
                raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Invalid file upload")
            file_bytes = await upload.read()
            if not file_bytes:
                raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Uploaded file is empty")
            new_type = DatasetType(dataset.type)
            if type_value:
                try:
                    new_type = DatasetType(str(type_value).strip().lower())
                except ValueError as exc:
                    raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Invalid dataset type") from exc
            if new_type == DatasetType.CSV:
                columns = _compute_csv_columns(file_bytes)
                sheet_name = None
            elif new_type == DatasetType.EXCEL:
                columns, sheet_name = _compute_excel_columns(
                    file_bytes,
                    sheet=str(sheet_param) if sheet_param else None,
                )
            else:
                raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Inline datasets must be updated via JSON")

            storage.remove_dataset(str(context.project.id), str(dataset.id))
            relative_path = storage.save_file(
                str(context.project.id),
                str(dataset.id),
                upload.filename or f"dataset.{new_type.value}",
                file_bytes,
            )
            source_payload: dict[str, Any] = {
                "path": relative_path,
                "filename": upload.filename or relative_path.rsplit("/", 1)[-1],
            }
            if sheet_name:
                source_payload["sheet"] = sheet_name
            dataset.type = new_type
            dataset.schema = _build_schema(columns)
            dataset.source = source_payload
        else:
            if type_value:
                desired_type = str(type_value).strip().lower()
                if desired_type != dataset.type.value:
                    raise http_exception(
                        status.HTTP_400_BAD_REQUEST,
                        ErrorCode.BAD_REQUEST,
                        "Upload a file to change dataset type",
                    )

        db.add(dataset)
        db.commit()
        db.refresh(dataset)
        return success_response(_serialize(dataset))

    raw_payload = await request.json()
    if not isinstance(raw_payload, dict):
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Invalid JSON payload")
    payload = DatasetUpdate.model_validate(raw_payload)
    updates = payload.model_dump(exclude_unset=True)

    if "name" in updates and updates["name"] is not None:
        dataset.name = _validate_dataset_name(updates["name"])

    desired_type = DatasetType(updates["type"]) if "type" in updates and updates["type"] is not None else dataset.type

    if desired_type != dataset.type and dataset.type != DatasetType.INLINE:
        raise http_exception(
            status.HTTP_400_BAD_REQUEST,
            ErrorCode.BAD_REQUEST,
            "File datasets can only be changed via file upload",
        )

    if desired_type == DatasetType.INLINE:
        source_payload = updates.get("source") or dataset.source or {}
        rows = source_payload.get("rows") if isinstance(source_payload, dict) else None
        if rows is None:
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Inline dataset requires rows")
        if any(not isinstance(row, dict) for row in rows):
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Each inline row must be an object")
        columns = _extract_columns_from_rows(rows)
        if not columns:
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Unable to infer dataset columns")
        dataset.type = DatasetType.INLINE
        dataset.schema = _build_schema(columns)
        dataset.source = {"rows": rows}
        storage.remove_dataset(str(context.project.id), str(dataset.id))
    else:
        if "source" in updates:
            raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, "Upload a file to update dataset contents")

    db.add(dataset)
    db.commit()
    db.refresh(dataset)
    return success_response(_serialize(dataset))


@router.delete("/{dataset_id}", response_model=ResponseEnvelope)
def delete_dataset(
    dataset_id: UUID,
    context: ProjectContext = Depends(require_project_member),
    db: Session = Depends(get_db),
) -> dict:
    dataset = _get_dataset(db, context.project.id, dataset_id)
    dataset.is_deleted = True
    db.add(dataset)
    db.commit()
    storage = _dataset_storage()
    storage.remove_dataset(str(context.project.id), str(dataset.id))
    return success_response({"id": dataset.id, "deleted": True})


@router.get("/{dataset_id}/preview", response_model=ResponseEnvelope)
def preview_dataset(
    dataset_id: UUID,
    limit: int = 20,
    context: ProjectContext = Depends(require_project_member),
    db: Session = Depends(get_db),
) -> dict:
    dataset = _get_dataset(db, context.project.id, dataset_id)
    try:
        rows = load_dataset_rows(dataset, limit=max(1, min(limit, 100)))
    except DatasetLoadError as exc:
        raise http_exception(status.HTTP_400_BAD_REQUEST, ErrorCode.BAD_REQUEST, str(exc)) from exc
    preview = DatasetPreview(dataset_id=dataset.id, rows=rows, total_rows=None)
    return success_response(preview.model_dump())
