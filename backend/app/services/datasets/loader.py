from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import get_settings
from app.models.dataset import Dataset, DatasetType


class DatasetLoadError(RuntimeError):
    """Raised when dataset rows cannot be loaded."""


def load_dataset_rows(dataset: Dataset, *, limit: int | None = None) -> list[dict[str, Any]]:
    if dataset.is_deleted:
        raise DatasetLoadError("Dataset has been deleted")

    if dataset.type == DatasetType.INLINE:
        return _load_inline_rows(dataset, limit=limit)
    if dataset.type == DatasetType.CSV:
        return _load_csv_rows(dataset, limit=limit)
    if dataset.type == DatasetType.EXCEL:
        return _load_excel_rows(dataset, limit=limit)
    raise DatasetLoadError(f"Unsupported dataset type: {dataset.type}")


def _load_inline_rows(dataset: Dataset, *, limit: int | None) -> list[dict[str, Any]]:
    source = dataset.source or {}
    rows = source.get("rows", []) if isinstance(source, dict) else []
    if not isinstance(rows, list):
        raise DatasetLoadError("Inline dataset rows must be a list")
    columns = _extract_columns(dataset)
    prepared: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            raise DatasetLoadError("Each inline dataset row must be an object")
        prepared.append(_normalize_row(row, columns))
        if limit is not None and len(prepared) >= limit:
            break
    return prepared


def _load_csv_rows(dataset: Dataset, *, limit: int | None) -> list[dict[str, Any]]:
    source = dataset.source or {}
    relative_path = source.get("path") if isinstance(source, dict) else None
    if not relative_path or not isinstance(relative_path, str):
        raise DatasetLoadError("CSV dataset is missing the file path")

    path = _resolve_storage_path(relative_path)
    if not path.exists():
        raise DatasetLoadError("CSV dataset file does not exist")

    columns = _extract_columns(dataset)
    prepared: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            prepared.append(_normalize_row(row, columns))
            if limit is not None and len(prepared) >= limit:
                break
    return prepared


def _load_excel_rows(dataset: Dataset, *, limit: int | None) -> list[dict[str, Any]]:
    source = dataset.source or {}
    relative_path = source.get("path") if isinstance(source, dict) else None
    sheet_name = source.get("sheet") if isinstance(source, dict) else None
    if not relative_path or not isinstance(relative_path, str):
        raise DatasetLoadError("Excel dataset is missing the file path")

    path = _resolve_storage_path(relative_path)
    if not path.exists():
        raise DatasetLoadError("Excel dataset file does not exist")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise DatasetLoadError(f"Sheet '{sheet_name}' not found in workbook")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration as exc:  # pragma: no cover - empty sheet
            raise DatasetLoadError("Excel dataset is empty") from exc
        header_names = [str(column).strip() if column is not None else "" for column in headers]
        if not any(header_names):
            raise DatasetLoadError("Excel dataset header row is empty")
        columns = _extract_columns(dataset) or [name for name in header_names if name]
        prepared: list[dict[str, Any]] = []
        for row_values in rows_iter:
            row_payload = {
                header_names[index]: row_values[index] if index < len(row_values) else None
                for index in range(len(header_names))
            }
            prepared.append(_normalize_row(row_payload, columns))
            if limit is not None and len(prepared) >= limit:
                break
        return prepared
    finally:
        workbook.close()


def _extract_columns(dataset: Dataset) -> list[str]:
    schema = dataset.schema or {}
    if isinstance(schema, dict):
        columns = schema.get("columns")
    else:
        columns = None
    if isinstance(columns, list):
        return [str(item) for item in columns if isinstance(item, (str, int))]
    return []


def _normalize_row(row: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    if not columns:
        return dict(row)
    normalized: dict[str, Any] = {}
    missing: list[str] = []
    for column in columns:
        normalized[column] = row.get(column)
        if column not in row:
            missing.append(column)
    if missing:
        raise DatasetLoadError(f"Row is missing required columns: {', '.join(missing)}")
    return normalized


def _resolve_storage_path(relative_path: str) -> Path:
    settings = get_settings()
    base_dir = Path(settings.dataset_storage_dir).expanduser().resolve()
    return base_dir.joinpath(relative_path)


__all__ = ["DatasetLoadError", "load_dataset_rows"]
